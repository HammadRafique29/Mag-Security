import os
import sys
import csv
import copy
import signal
import base64
import requests
import platform
from typing import List
from io import StringIO
from typing import List
from getpass import getpass
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from cryptography.hazmat.primitives import padding, hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend

IS_WINDOWS_OS = platform.system() == "Windows"
ENCRYPTED_OUTPUT = os.path.join(os.getcwd(), "encrypted_outputs")
DECRYPTED_OUTPUT = os.path.join(os.getcwd(), "decrypted_outputs")

if not os.path.exists(ENCRYPTED_OUTPUT): os.mkdir(ENCRYPTED_OUTPUT)
if not os.path.exists(DECRYPTED_OUTPUT): os.mkdir(DECRYPTED_OUTPUT)

info = """
#############################################################################################################
##                                                                                                         ##
##   888b     d888                    .d8888b.                                     d8b 888                 ##
##   8888b   d8888                   d88P  Y88b                                    Y8P 888                 ##
##   88888b.d88888                   Y88b.                                             888                 ##
##   888Y88888P888  8888b.   .d88b.   "Y888b.    .d88b.   .d8888b 888  888 888d888 888 888888 888  888     ##
##   888 Y888P 888     "88b d88P"88b     "Y88b. d8P  Y8b d88P"    888  888 888P"   888 888    888  888     ##
##   888  Y8P  888 .d888888 888  888       "888 88888888 888      888  888 888     888 888    888  888     ##
##   888   "   888 888  888 Y88b 888 Y88b  d88P Y8b.     Y88b.    Y88b 888 888     888 Y88b.  Y88b 888     ##
##   888       888 "Y888888  "Y88888  "Y8888P"   "Y8888   "Y8888P  "Y88888 888     888  "Y888  "Y88888     ##
##                               888                                                               888     ##
##                          Y8b d88P                                                          Y8b d88P     ##
##                           "Y88P"                                                            "Y88P"      ##
##                                                                                                         ##
#############################################################################################################                     
"""

outro = """
┳┓┓┏┏┓╻  ┳┓    ╹   ┏                                        ┓┏       ┓             ┓               ┓        
┣┫┗┫┣ ┃  ┃┃┏┓┏┓ ╋  ╋┏┓┏┓┏┓┏┓╋  ╋┏┓  ┏┓┏┓┏┓╋┏┓┏╋  ┓┏┏┓┓┏┏┓┏┏┓┃╋  ┏┓┏┓┏┫  ┓┏┏┓┓┏┏┓  ┏┫┏┓╋┏┓  ┏┓┓┏╋  ╋┣┓┏┓┏┓┏┓ 
┻┛┗┛┗┛•  ┻┛┗┛┛┗ ┗  ┛┗┛┛ ┗┫┗ ┗  ┗┗┛  ┣┛┛ ┗┛┗┗ ┗┗  ┗┫┗┛┗┻┛ ┛┗ ┗┛  ┗┻┛┗┗┻  ┗┫┗┛┗┻┛   ┗┻┗┻┗┗┻  ┗┛┗┻┗  ┗┛┗┗ ┛ ┗ •
                         ┛          ┛             ┛                      ┛                                  
"""

backend = default_backend()


# FUNCTION: CLEAR SCREEN
def clear_screen():
    if IS_WINDOWS_OS: os.system("cls")
    else: os.system("clear")
    print(info)


# FUNCTION: CHECK STR IS BASE64
def is_base64(s):
    try:
        base64.b64decode(s, validate=True)
        return True
    except Exception:return False
    

# FUNCTION: CREATE KEY BASED ON PASSWORD
def derive_key(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100_000,
        backend=backend
    )
    return kdf.derive(password.encode())



# FUNCTION: TO ENCRYPT THE DATA
def encrypt(message: str, password: str):
    salt = os.urandom(16)
    key = derive_key(password, salt)
    iv = os.urandom(16)

    padder = padding.PKCS7(128).padder()
    padded = padder.update(message.encode()) + padder.finalize()

    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=backend)
    encryptor = cipher.encryptor()
    ciphertext = encryptor.update(padded) + encryptor.finalize()

    result = base64.b64encode(salt + iv + ciphertext).decode()
    return result



# FUNCTION: TO DECRYPT ENCRYPTED DATA
def decrypt(token: str, password: str):
    raw = base64.b64decode(token)
    salt, iv, ciphertext = raw[:16], raw[16:32], raw[32:]
    key = derive_key(password, salt)

    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=backend)
    decryptor = cipher.decryptor()
    padded_plaintext = decryptor.update(ciphertext) + decryptor.finalize()

    unpadder = padding.PKCS7(128).unpadder()
    plaintext = unpadder.update(padded_plaintext) + unpadder.finalize()
    return plaintext.decode()



# FUNCTION: SAVE DATA IN LOCAL XLSX SHEET
def save_to_excel(data: List[List[str]], filename: str):
    """
    Saves a 2D list to an Excel (.xlsx) file.
    Args:
        data (List[List[str]]): The 2D list to write into the Excel file.
        filename (str): The name of the Excel file to save (should end with .xlsx).
    """
    wb = Workbook()
    ws = wb.active
    for row in data: ws.append(row)
    wb.save(filename)
    return filename



# FUNCTION: READ DATA FROM SHEET FILE (GOOGLE SHEET)
def download_google_sheet(sheet_url: str, remove_header: bool = False) -> List[List[str]]:
    """
    Downloads a Google Sheet in CSV format and returns its content as a 2D list.
    Args:
        sheet_url (str): The public URL of the Google Sheet.
        remove_header (bool): If True, removes the first row (assumed to be header).
    Returns: List[List[str]]: 2D list representing rows and columns of the sheet.
    """
    try:
        try: sheet_id = sheet_url.split("/d/")[1].split("/")[0]
        except IndexError: raise ValueError("Invalid Google Sheet URL format.")
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        response = requests.get(csv_url)
        if response.status_code != 200: raise Exception("Failed to download the Google Sheet. Check access permissions.")
        data = list(csv.reader(StringIO(response.text)))
        if remove_header and data: data = data[1:]
        return data
    except Exception as e: raise Exception(str(e))



# FUNCTION: READ DATA FROM SHEET FILE (LOCAL)
def read_excel_to_list(filename: str) -> List[List[str]]:
    """
    Reads an Excel (.xlsx) file and returns its content as a 2D list.
    Args: filename (str): The path to the Excel file.
    Returns: List[List[str]]: A list of rows, where each row is a list of cell values as strings.
    """
    data = []
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True): data.append([str(cell) if cell is not None else '' for cell in row])
    return data



# FUNCTION: TO PRINT SHEETS DATA
def print_table(data, min_width=5, max_width=20):
    for row in data:
        formatted_row = []
        for cell in row:
            cell_str = str(cell)
            if len(cell_str) > max_width: cell_str = cell_str[:max_width - 3] + '...'
            formatted_cell = cell_str.ljust(min_width)
            formatted_row.append(formatted_cell)
        print(' | '.join(formatted_row))



# FUNCTION TO HANDLE UNEXPECTED INTERRUPT
def handle_exit(signum, frame):
    print("\n\n----- Program interrupted. Exiting safely. Stay secure out there! -----")
    sys.exit(0)

def setup_ctrl_c_handler():
    signal.signal(signal.SIGINT, handle_exit)





if __name__ == "__main__":

    setup_ctrl_c_handler()
    
    
    while True:
        clear_screen()
        print("-- 🔐 HINT: Password Is Required To Encrypt / Decrypt Your Data.")
        print("-- 🧠 Please Remember This Password, As It's The Only Way To Access Your Encrypted Data later.")
        print("-- 🙈 Note: The Password You Type Will Be Hidden For Security (nothing will appear as you type).")

        password = getpass("\nEnter password: ").strip()
        if password: break
        print("--> Password Cannot Be Empty!")
        input()


    while True:
        clear_screen()
        print("\n################# Personal Security Program: \n")
        print("1. Start Encryption")
        print("2. Start Decryption")
        print("3. Start Password Auth")
        print("4. Exit...")

        choice = input("\nChoice (1-4): ").strip().lower()

        if choice == '1':
            
            while True:
                clear_screen()
                print("\n################# Encryption: \n")
                print("1. Encrypt Single Message")
                print("2. Encrypt Sheet Data")
                print("3. Go Back...")
                case = input("\nChoice (1-3): ")

                match case:

                    case "1":
                        msg = input("\nEnter message to encrypt: ")
                        try: print(f"--> Encrypted: {encrypt(msg, password)}") 
                        except Exception as e: print("Error: ", e)
                        input()

                    case "2":
                        sheet_loc = input("\nEnter Goole Sheet URL: ")
                        try: 
                            if "http" in sheet_loc: data = download_google_sheet(sheet_loc, remove_header=True)
                            else: data = read_excel_to_list(sheet_loc)
                            print("\n#### Sheet Data:")
                            print_table(data)

                            temp = copy.deepcopy(data)
                            for Rindex, row in enumerate(temp):
                                for Cindex, col in enumerate(row):
                                    data[Rindex][Cindex] = encrypt(col, password)
                            
                            print("\n#### Encrypted Sheet Data:")
                            print_table(data)
                            if input("\nWANT TO SAVE DATA (Y/N): ").lower() == 'y':
                                cur_time = datetime.now()
                                fileName = os.path.join(ENCRYPTED_OUTPUT, f"encrypted_output_{cur_time.hour}_{cur_time.minute}_{cur_time.second}.xlsx")
                                output_ = save_to_excel(data, fileName)
                                print(f"--> Saved to {output_}")

                        except Exception as e: print("Error: ", e)
                        input()

                    case "3": break


        elif choice == '2':
            while True:
                clear_screen()
                print("\n################# Encryption: \n")
                print("1. Decrypt Single Message")
                print("2. Decrypt Sheet Data")
                print("3. Go Back...")
                case = input("\nChoice (1-3): ")

                match case:
                    
                    case "1":
                        token = input("\nEnter encrypted text: ")
                        try: 
                            if is_base64(token): print(f"--> Decrypted: {decrypt(token, password)}")
                            else: raise Exception("Token Need To Be Encrypted One..") 
                        except Exception as e: print("--> Error: ", e)
                        input()

                    case "2":
                        sheet_loc = input("\nEnter Goole Sheet URL or XLSX Path: ")
                        try: 
                            if "http" in sheet_loc: data = download_google_sheet(sheet_loc, remove_header=True)
                            else: data = read_excel_to_list(sheet_loc)
                            print("\n#### Sheet Data:")
                            print_table(data)
                            temp = copy.deepcopy(data)
                            for Rindex, row in enumerate(temp):
                                for Cindex, col in enumerate(row):
                                    if is_base64(col): data[Rindex][Cindex] = decrypt(col, password)
                                    else: data[Rindex][Cindex] = col
                            
                            print("\n#### Decrypted Sheet Data:")
                            print_table(data)
                            if input("\nWANT TO SAVE DATA (Y/N): ").lower() == 'y':
                                cur_time = datetime.now()
                                fileName = os.path.join(DECRYPTED_OUTPUT, f"decrypted_output_{cur_time.hour}_{cur_time.minute}_{cur_time.second}.xlsx")
                                output_ = save_to_excel(data, fileName)
                                print(f"--> Saved to {output_}")

                        except Exception as e: print("Error: ", e)
                        input()

                    case "3": break


        elif choice == '3': 

            while True:
                clear_screen()
                print("\n################# Encryption: \n")
                print("1. Verify Password (Match With Old)")
                print("2. Update Password (Change Password)")
                print("3. Go Back...")
                case = input("\nChoice (1-3): ")

                match case:
                    
                    case "1":
                        temp_pass = getpass("\nEnter your password: ").strip()
                        if temp_pass == password: print("--> Password Matched!")
                        else:  print("--> Password Not Matched!")
                        input()

                    case "2":
                        temp_pass = getpass("\nEnter new password: ").strip()
                        if temp_pass: 
                            password = temp_pass
                            hidden = '*' * (len(password) - 2) if len(password) > 2 else ''
                            print(f"--> Password Changed {password[0]}{hidden}{password[-1]}!")
                        else:  print("--> Password Cannot Be Empty! Try Again...")
                        input()
                    
                    case '3': break


        elif choice == '4': 
            print("\n", outro)
            break

        else: 
            print("Invalid choice.")
            continue
